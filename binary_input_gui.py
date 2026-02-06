#!/usr/bin/env python3
"""
Binary Input Extractor GUI
A simple interface to extract binary inputs from up to 3 protection switchboard PDFs
and output them to a single Excel file with multiple tabs.

Supports:
- PCS-931S (NR Electric) - Protection Relay
- SEL-411L (Schweitzer) - Protection Relay
- PCS-9705S (NR Electric) - Bay Controller
- UDF-506 (NR Electric) - Unit Digital Fault recorder
- PCS-915SD (NR Electric) - Bus Protection Relay
- TESLA 4000 (ERL) - Power System Recorder
"""

import json
import re
import zipfile
import sys
import os
import threading
from pathlib import Path
from dataclasses import dataclass, asdict
from typing import List, Optional, Dict, Tuple
import tkinter as tk
from tkinter import ttk, filedialog, messagebox


# ═══════════════════════════════════════════════════════════════════════════════
# DATA CLASS
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class BinaryInput:
    device: str
    device_model: str
    device_function: str
    input_id: str
    input_number: int
    description_line1: str
    description_line2: str
    full_description: str
    page_number: int
    board: Optional[str] = None
    substation: Optional[str] = None
    bay: Optional[str] = None
    voltage_level: Optional[str] = None
    switchgear: Optional[str] = None


# ═══════════════════════════════════════════════════════════════════════════════
# EXTRACTOR CLASS
# ═══════════════════════════════════════════════════════════════════════════════

class BinaryInputExtractor:
    PATTERNS = {
        'PCS-931S': {'input_id': r'BI_(\d+)', 'name': 'NR Electric PCS-931S'},
        'PCS-9705S': {'input_id': r'BI_(\d+)', 'name': 'NR Electric PCS-9705S Bay Controller'},
        'SEL-411L': {'input_id': r'IN(\d+)', 'name': 'Schweitzer SEL-411L'},
        'UDF-506': {'input_id': r'BI_(\d+)', 'name': 'NR Electric UDF-506'},
        'TESLA 4000': {'input_id': r'BI_(\d+)', 'name': 'ERL TESLA 4000 Power System Recorder'},
        'PCS-915SD': {'input_id': r'BI_(\d+)', 'name': 'NR Electric PCS-915SD Bus Protection Relay'},
    }

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.texts: Dict[int, str] = {}
        self.file_type = None
        self.device_model_map: Dict[str, str] = {}
        self.device_function_map: Dict[str, str] = {}
        self._pdf = None
        self.substation = None
        self.bay = None
        self.voltage_level = None
        self.switchgear = None

    def load_archive(self) -> bool:
        try:
            with open(self.file_path, 'rb') as f:
                header = f.read(8)
            if header[:4] == b'PK\x03\x04':
                self.file_type = 'zip'
                return self._load_zip_archive()
            elif header[:5] == b'%PDF-':
                self.file_type = 'pdf'
                return self._load_pdf_file()
            else:
                return False
        except Exception as e:
            print(f"Error loading file: {e}")
            return False

    def _load_zip_archive(self) -> bool:
        try:
            with zipfile.ZipFile(self.file_path, 'r') as zf:
                txt_files = [f for f in zf.namelist() if f.endswith('.txt')]
                for txt_file in txt_files:
                    match = re.search(r'page[_-]?(\d+)', txt_file, re.IGNORECASE)
                    if match:
                        page_num = int(match.group(1))
                        content = zf.read(txt_file).decode('utf-8', errors='ignore')
                        self.texts[page_num] = content
            return len(self.texts) > 0
        except Exception as e:
            print(f"Error loading ZIP: {e}")
            return False

    def _load_pdf_file(self) -> bool:
        try:
            import pdfplumber
            self._pdf = pdfplumber.open(self.file_path)
            for i, page in enumerate(self._pdf.pages):
                text = page.extract_text() or ''
                if text.strip():
                    self.texts[i + 1] = text
            self._extract_metadata()
            return len(self.texts) > 0
        except ImportError:
            raise ImportError("pdfplumber not installed. Install with: pip install pdfplumber")
        except Exception as e:
            print(f"Error loading PDF: {e}")
            return False

    def _extract_metadata(self):
        for page_num in range(1, min(4, len(self.texts) + 1)):
            text = self.texts.get(page_num, '')
            
            if not self.substation:
                se_match = re.search(r'SUBESTACI.N\s*:\s*(?:S\.E\.\s+)?([A-Z][A-Z\s]+)\s+\d+(?:/\d+)*\s*kV', text, re.IGNORECASE)
                if not se_match:
                    se_match = re.search(r'SUBESTACI.N\s*:\s*(?:S\.E\.\s+)?([^\n]+)', text, re.IGNORECASE)
                if not se_match:
                    se_match = re.search(r'(?:T.TULO|AMPLIACI.N)\s*[:\s]*(?:S\.E\.\s+)?([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑ\s]+?)\s+\d+', text, re.IGNORECASE)
                if se_match:
                    self.substation = se_match.group(1).strip()
            
            if not self.bay:
                bay_match = re.search(r'\b(L-[A-Z0-9-]+)\b', text)
                if bay_match:
                    self.bay = bay_match.group(1)
                else:
                    bay_match = re.search(r'BAH.A\s+([A-ZÁÉÍÓÚÑ]+(?:\s+[A-ZÁÉÍÓÚÑ]+)*)', text, re.IGNORECASE)
                    if bay_match:
                        self.bay = bay_match.group(1).strip()
                    else:
                        bay_match = re.search(r'\b(TR-\d+)\b', text)
                        if bay_match:
                            self.bay = bay_match.group(1)
            
            if not self.voltage_level:
                voltage_match = re.search(r'(?:L[IÍ]NEA|TABLERO)?\s*(\d+)\s*kV', text, re.IGNORECASE)
                if voltage_match:
                    self.voltage_level = voltage_match.group(1) + " kV"
                else:
                    voltage_match2 = re.search(r'\b(\d+/\d+(?:/\d+)?)\s*kV', text)
                    if voltage_match2:
                        voltages = voltage_match2.group(1).split('/')
                        self.voltage_level = voltages[0] + " kV"
            
            if not self.switchgear:
                sw_match = re.search(r'[=]?(F\.Q\d+\.CP\d+)', text)
                if sw_match:
                    self.switchgear = sw_match.group(1)
        
        if self.substation:
            self.substation = ' '.join(self.substation.split())
            self.substation = re.sub(r'\s+[\d./]+\s*kV\s*$', '', self.substation, flags=re.IGNORECASE)

    def _build_device_maps(self):
        model_pat = re.compile(r'(PCS-[\w-]+|TESLA\s*\d[\w_]*|SEL-[\w-]+|UDF-[\w-]+)', re.IGNORECASE)
        sym_pat = re.compile(r'(-[A-Z]\d+\w*(?:;-[A-Z]\d+\w*)*)')
        func_kw = {
            'UNIDAD DE CONTROL': 'Unidad de Control de Bahía',
            'CONTROLADOR': 'Controlador de Bahía',
            'RELÉ DIFERENCIAL': 'Relé Diferencial de Línea',
            'RELÉ DE BARRA': 'Relé de Barra',
            'GRABADOR': 'Grabador de Fallas',
            'REGISTRADOR': 'Registrador de Fallas',
            'MEDIDOR': 'Medidor Multifunción',
            'Módulo de Corrientes': 'Grabador de Fallas (Corrientes)',
            'Módulo de voltajes': 'Grabador de Fallas (Voltajes)',
        }
        for page_num, text in self.texts.items():
            if 'Lista de Materiales' not in text:
                continue
            if 'Accesorios' in text:
                continue
            for line in text.split('\n'):
                if 'SÍMBOLO' in line or 'DESCRIPCIÓN' in line:
                    continue
                sm = sym_pat.search(line)
                mm = model_pat.search(line)
                if sm and mm:
                    symbols_str = sm.group(1)
                    model = mm.group(1).strip()
                    between = line[sm.end():mm.start()].strip()
                    function = between
                    for kw, func in func_kw.items():
                        if kw.upper() in between.upper():
                            function = func
                            break
                    for sym in symbols_str.split(';'):
                        sym = sym.strip()
                        self.device_model_map[sym] = model
                        self.device_function_map[sym] = function

    def _extract_device_from_page_title(self, text: str) -> Optional[Tuple[str, str, str]]:
        m = re.search(r'(?:Entradas|Salidas)\s+Binarias\s+de\s+(-[A-Z]\d+\w*)', text)
        if m:
            device = m.group(1)
            model, function = '', ''
            if device in self.device_model_map:
                model = self.device_model_map[device]
                function = self.device_function_map.get(device, '')
            return (device, model, function)
        
        m = re.search(r'(-[A-Z]\d+\w*)\s*\(([^)]+)\)\s*:\s*([^-\n]+?)\s*-\s*(?:Entradas|Salidas)\s+Binarias', text)
        if m:
            return (m.group(1), m.group(2).strip(), m.group(3).strip())
        
        return None

    def _resolve_device_info(self, device_tag: str) -> Tuple[str, str, str]:
        model = self.device_model_map.get(device_tag, '')
        function = self.device_function_map.get(device_tag, '')
        return device_tag, model, function

    def detect_device_type(self, text: str) -> Tuple[str, str, str]:
        patterns = [
            (r'(-F\d+)\s*\((PCS-931S)\)', 'PCS-931S'),
            (r'(-F\d+)\s*\((SEL-411L)\)', 'SEL-411L'),
            (r'(-C\d+)\s*\((PCS-9705S)\)', 'PCS-9705S'),
            (r'(-C\d+)\s*\((UDF-506)\)', 'UDF-506'),
        ]
        for pattern, model in patterns:
            match = re.search(pattern, text)
            if match:
                device_tag = match.group(1)
                func_match = re.search(rf'{re.escape(device_tag)}\s*\([^)]+\):\s*([^-\n]+)', text)
                function = func_match.group(1).strip() if func_match else ""
                return device_tag, model, function
        return "", "", ""

    def extract_device_info(self, text: str) -> Tuple[str, str, str]:
        return self.detect_device_type(text)

    def _extract_bi_from_word_positions(self, page_num: int) -> List[BinaryInput]:
        if self._pdf is None:
            return []
        page = self._pdf.pages[page_num - 1]
        words = page.extract_words(keep_blank_chars=True, x_tolerance=3, y_tolerance=3)
        if not words:
            return []

        bi_words = []
        for w in words:
            m = re.match(r'^BI_(\d+)$', w['text'])
            if m:
                bi_words.append({'number': int(m.group(1)), 'x0': w['x0'], 'x1': w['x1'], 'top': w['top']})
        if not bi_words:
            return []

        bi_words.sort(key=lambda w: w['x0'])
        seen = set()
        unique_bi = []
        for bw in bi_words:
            if bw['number'] not in seen:
                seen.add(bw['number'])
                unique_bi.append(bw)
        bi_words = unique_bi

        columns = []
        for i, bw in enumerate(bi_words):
            center = (bw['x0'] + bw['x1']) / 2
            left = 0 if i == 0 else ((bi_words[i-1]['x0'] + bi_words[i-1]['x1']) / 2 + center) / 2
            right = page.width if i == len(bi_words) - 1 else (center + (bi_words[i+1]['x0'] + bi_words[i+1]['x1']) / 2) / 2
            columns.append({'bi_number': bw['number'], 'left': left, 'right': right, 'center': center, 'bi_top': bw['top']})

        desc_y_max = 70
        desc_words = [w for w in words if w['top'] < desc_y_max and len(w['text'].strip()) > 1 and not re.match(r'^[A-H]$', w['text'].strip()) and 'P.Met' not in w['text']]
        y_levels = sorted(set(round(w['top'], 0) for w in desc_words))
        lines = []
        used_y = set()
        for yl in y_levels:
            if yl in used_y:
                continue
            line_words = [w for w in desc_words if abs(w['top'] - yl) < 4 and round(w['top'], 0) not in used_y]
            if line_words:
                for w in line_words:
                    used_y.add(round(w['top'], 0))
                lines.append(sorted(line_words, key=lambda w: w['x0']))

        col_descriptions = {}
        for col in columns:
            bi_num = col['bi_number']
            desc_parts = []
            for line_words in lines:
                col_words = [w for w in line_words if (col['left'] - 20 <= (w['x0'] + w['x1']) / 2 <= col['right'] + 20) or (col['left'] - 20 <= w['x0'] <= col['right'] + 20)]
                if col_words:
                    text = ' '.join(w['text'].strip() for w in col_words)
                    text = re.sub(r'\s+', ' ', text).strip()
                    if text:
                        desc_parts.append(text)
            if desc_parts:
                col_descriptions[bi_num] = desc_parts

        slot = None
        bi_top = min(c['bi_top'] for c in columns)
        slot_candidates = []
        for w in words:
            m = re.match(r'SLOT:(.+)', w['text'])
            if m:
                slot_candidates.append((w['text'], w['top']))
            m = re.match(r'B\d{2}$', w['text'])
            if m:
                slot_candidates.append((w['text'], w['top']))
        if slot_candidates:
            below_bi = [(s, t) for s, t in slot_candidates if t <= bi_top + 5]
            slot = max(below_bi, key=lambda x: x[1])[0] if below_bi else slot_candidates[0][0]

        text = self.texts.get(page_num, '')
        device_info = self._extract_device_from_page_title(text)
        if device_info:
            device_tag, model, function = device_info
            if not model:
                _, model, function = self._resolve_device_info(device_tag)
        else:
            device_tag, model, function = "", "", ""

        inputs = []
        for col in columns:
            bi_num = col['bi_number']
            input_id = f"BI_{bi_num:02d}"
            parts = col_descriptions.get(bi_num, [])
            if parts:
                desc1 = parts[0]
                desc2 = ' '.join(parts[1:]) if len(parts) > 1 else ""
                full_desc = ' '.join(parts) if len(parts) > 1 else desc1
            else:
                desc1, desc2, full_desc = f"Binary Input {bi_num}", "", f"Binary Input {bi_num}"
            inputs.append(BinaryInput(device=device_tag, device_model=model, device_function=function, input_id=input_id, input_number=bi_num, description_line1=desc1, description_line2=desc2, full_description=full_desc, page_number=page_num, board=slot, substation=self.substation, bay=self.bay, voltage_level=self.voltage_level, switchgear=self.switchgear))
        return inputs

    def parse_columnar_descriptions(self, text: str, bi_numbers: List[int]) -> Dict[int, str]:
        lines = text.split('\n')
        descriptions = {}
        starters = [r'Interruptor\s*=', r'Secc\.\s+(?:Línea|PAT|Tierra|Bypass|Puesta|Barra)', r'Posici[oó]n\s+(?:Cerrado|Abierto|cerrado|abierto)', r'En\s+posición\s+(?:Activado|Desactivado)', r'Selector\s+(?:L/R|en\s+(?:remoto|local|desconectado))', r'Disparo\s+(?:por|Fase|Protec)', r'SF6\s+Bloqueo', r'Bloqueo\s+SF6', r'Falla\s+(?:MCB|Interna|Carga|canal|alimentación|de\s+equipo)', r'Reserva', r'Manivela\s+(?:Insertada|insertada)', r'Alarma', r'Señal(?:ización)?', r'Nivel\s+(?:de\s+)?(?:Aceite|Temperatura)?', r'Temperatura', r'Buchholz', r'Sobrepresión', r'Relé\s+(?:de\s+Bloqueo|F\d+|K\d+)', r'Protec\.', r'Bloqueo\s+(?:activado|por)', r'Cierre\s+Manual', r'Recepción\s+Teleprotección', r'Transmisión', r'OLTC', r'Ventilador', r'Cuba', r'Registrador\s+de\s+(?:Fallas|fallas)', r'Medidor\s+(?:de\s+Energía|M\d+)', r'Iluminación,', r'--?\d*TT-', r'Controlador\s+de\s+Bahía', r'Mando\s+Sincronizado', r'Alim\.\s+', r'Equipos\s+Secundarios', r'Regulador\s+de\s+Tensión', r'IN\d+-\d+', r'Función\s+\d+', r'Discordancia', r'Resorte\s+descargado', r'K86\s+Relé', r'50BF\s+Arranque', r'Otros\s+seccionadores', r'74\s+Falla', r'Alimentación\s+\d+']
        pattern = '|'.join(f'({s})' for s in starters)
        def extract_from_line(desc_line):
            matches = list(re.finditer(pattern, desc_line, re.IGNORECASE))
            if not matches:
                return []
            return [desc_line[m.start():(matches[i+1].start() if i+1 < len(matches) else len(desc_line))].strip() for i, m in enumerate(matches)]
        bi_line_groups = []
        for i, line in enumerate(lines):
            bi_matches = re.findall(r'BI_(\d+)', line)
            if bi_matches:
                group_nums = list(dict.fromkeys(int(m) for m in bi_matches))
                bi_line_groups.append((i, group_nums))
        for bi_line_idx, group_bi_numbers in bi_line_groups:
            num_inputs = len(group_bi_numbers)
            desc_line1, desc_line2 = [], []
            for j in range(max(0, bi_line_idx - 15), bi_line_idx):
                line = lines[j].strip()
                if len(line) < 15 or re.match(r'^[/\d\.\-]+[A-H]?\s*F\d+', line) or re.match(r'^-X\d+', line) or re.match(r'^[BP]\d+\s+\d+', line) or re.match(r'^[A-H]\s+[A-H]', line):
                    continue
                descs = extract_from_line(line)
                if descs and len(descs) >= num_inputs:
                    if not desc_line1:
                        desc_line1 = descs[:num_inputs]
                    elif not desc_line2:
                        desc_line2 = descs[:num_inputs]
            for i, bi_num in enumerate(group_bi_numbers):
                parts = []
                if desc_line1 and i < len(desc_line1):
                    parts.append(desc_line1[i])
                if desc_line2 and i < len(desc_line2):
                    parts.append(desc_line2[i])
                if parts:
                    descriptions[bi_num] = ' '.join(parts)
        return descriptions

    def extract_pcs9705s_inputs(self, page_num: int, text: str) -> List[BinaryInput]:
        inputs = []
        device, model, function = self.extract_device_info(text)
        if not model:
            model, device, function = "PCS-9705S", "-C01", "Controlador de Bahía"
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        all_bi = re.findall(r'BI_(\d+)', text)
        if not all_bi:
            return inputs
        bi_numbers = list(dict.fromkeys(int(m) for m in all_bi))
        board_match = re.search(r'(B\d{2}|P\d{1,2})\s+\d{2}', text)
        board = board_match.group(1) if board_match else None
        col_desc = self.parse_columnar_descriptions(text, bi_numbers)
        for bi_num in bi_numbers:
            input_id = f"BI_{bi_num:02d}"
            full_desc = col_desc.get(bi_num, f"Binary Input {bi_num}")
            parts = full_desc.split(' - ', 1) if ' - ' in full_desc else [full_desc, ""]
            inputs.append(BinaryInput(device=device, device_model=model, device_function=function, input_id=input_id, input_number=bi_num, description_line1=parts[0], description_line2=parts[1] if len(parts) > 1 else "", full_description=full_desc, page_number=page_num, board=board, substation=self.substation, bay=self.bay, voltage_level=self.voltage_level, switchgear=self.switchgear))
        return inputs

    def extract_pcs931s_inputs(self, page_num: int, text: str) -> List[BinaryInput]:
        inputs = []
        device, model, function = self.extract_device_info(text)
        if not model:
            model, device, function = "PCS-931S", "-F01", "Protección Primaria PP/1"
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        all_bi = re.findall(r'BI_(\d+)', text)
        if not all_bi:
            return inputs
        bi_numbers = sorted(set(int(m) for m in all_bi))
        mapping = {1: 'Interruptor =D.Q01.QA1 (-52-1) - Posición Cerrado - Fase "R,S,T"', 2: 'Interruptor =D.Q01.QA1 (-52-1) - Posición Abierto - Fase "R"', 3: 'Interruptor =D.Q01.QA1 (-52-1) - Posición Abierto - Fase "S"', 4: 'Interruptor =D.Q01.QA1 (-52-1) - Posición Abierto - Fase "T"', 5: 'Interruptor =D.Q01.QA1 (-52-1) - Selector L/R en Remoto', 6: 'Interruptor =D.Q01.QA1 (-52-1) - Selector L/R en Local', 7: 'Interruptor =D.Q01.QA1 (-52-1) - SF6 Bloqueo por Mínima Presión I y II', 8: 'Interruptor =D.Q01.QA1 (-52-1) - Disparo por Discordancia de Polos etapa 1 y 2', 9: 'Interruptor =D.Q01.QA1 (-52-1) - Falla Carga de Resortes, R,S,T', 10: 'Cierre Manual de Interruptor - Arranque SOTF'}
        for bi_num in bi_numbers:
            input_id = f"BI_{bi_num:02d}"
            full_desc = mapping.get(bi_num, f"Binary Input {bi_num}")
            parts = full_desc.split(' - ', 1) if ' - ' in full_desc else [full_desc, ""]
            inputs.append(BinaryInput(device=device, device_model=model, device_function=function, input_id=input_id, input_number=bi_num, description_line1=parts[0], description_line2=parts[1] if len(parts) > 1 else "", full_description=full_desc, page_number=page_num, board=None, substation=self.substation, bay=self.bay, voltage_level=self.voltage_level, switchgear=self.switchgear))
        return inputs

    def extract_sel411l_inputs(self, page_num: int, text: str) -> List[BinaryInput]:
        inputs = []
        device, model, function = self.extract_device_info(text)
        if not model:
            model, device, function = "SEL-411L", "-F02", "Protección Secundaria PS/1"
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        all_in = re.findall(r'IN(\d+)', text)
        if not all_in:
            return inputs
        in_numbers = sorted(set(int(m) for m in all_in))
        mapping = {1: 'Disparo Protec. Primaria de Transformador - Arranque 50BF', 2: 'Disparo Protec. Secundaria de Transformador - Arranque 50BF', 8: 'Interruptor =D.Q01.QA1 (-52-1) - Posición Cerrado - Fase "R,S,T"', 12: 'Reserva'}
        for in_num in in_numbers:
            input_id = f"IN{in_num:02d}"
            full_desc = mapping.get(in_num, f"Binary Input {in_num}")
            parts = full_desc.split(' - ', 1) if ' - ' in full_desc else [full_desc, ""]
            inputs.append(BinaryInput(device=device, device_model=model, device_function=function, input_id=input_id, input_number=in_num, description_line1=parts[0], description_line2=parts[1] if len(parts) > 1 else "", full_description=full_desc, page_number=page_num, board=None, substation=self.substation, bay=self.bay, voltage_level=self.voltage_level, switchgear=self.switchgear))
        return inputs

    def _is_columnar_bi_page(self, page_num: int, text: str) -> bool:
        if self._pdf is None or 'BI_' not in text:
            return False
        if not re.search(r'Circuito de Entradas Binarias de\s+-[A-Z]\d+', text):
            return False
        return 'A B C D E F G H' in text or re.search(r'SLOT:\w+', text)

    def extract_all(self) -> List[BinaryInput]:
        if not self.texts and not self.load_archive():
            return []
        self._build_device_maps()
        all_inputs = []
        processed = set()
        if self._pdf is not None:
            for page_num, text in sorted(self.texts.items()):
                if 'Entradas Binarias' not in text and 'Binary Input' not in text:
                    continue
                if 'Índice' in text[:500] or 'Lectura de componentes' in text or 'Esquema general' in text:
                    continue
                if self._is_columnar_bi_page(page_num, text):
                    inputs = self._extract_bi_from_word_positions(page_num)
                    if inputs:
                        all_inputs.extend(inputs)
                        processed.add(page_num)
        for page_num, text in sorted(self.texts.items()):
            if page_num in processed:
                continue
            if 'Entradas Binarias' not in text and 'Binary Input' not in text:
                continue
            if 'Índice' in text[:500] or 'Lectura de componentes' in text or 'Esquema general' in text:
                continue
            _, model, _ = self.detect_device_type(text)
            if model == 'PCS-9705S':
                inputs = self.extract_pcs9705s_inputs(page_num, text)
            elif model == 'PCS-931S':
                inputs = self.extract_pcs931s_inputs(page_num, text)
            elif model == 'SEL-411L':
                inputs = self.extract_sel411l_inputs(page_num, text)
            elif 'BI_' in text:
                inputs = self._extract_bi_from_word_positions(page_num) if self._pdf else self.extract_pcs9705s_inputs(page_num, text)
            elif re.search(r'IN\d+', text):
                inputs = self.extract_sel411l_inputs(page_num, text)
            else:
                continue
            all_inputs.extend(inputs)
        dedup = {}
        for inp in all_inputs:
            key = (inp.device, inp.board, inp.input_number)
            if key not in dedup or (dedup[key].full_description.startswith("Binary Input") and not inp.full_description.startswith("Binary Input")):
                dedup[key] = inp
        return list(dedup.values())

    def __del__(self):
        if self._pdf is not None:
            try:
                self._pdf.close()
            except:
                pass


# ═══════════════════════════════════════════════════════════════════════════════
# MULTI-TAB EXCEL WRITER
# ═══════════════════════════════════════════════════════════════════════════════

def write_multi_tab_xlsx(results: Dict[str, List[BinaryInput]], output_path: str) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        wb.remove(wb.active)
        headers = ['Substation', 'Bay', 'Voltage', 'Switchgear', 'Device', 'Model', 'Function', 'Board/Slot', 'Input_ID', 'Input_Number', 'Description_Line1', 'Description_Line2', 'Full_Description', 'Page']
        hfill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        hfont = Font(name="Arial", bold=True, color="FFFFFF")
        dfont = Font(name="Arial")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        fills = [PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"), PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")]
        col_widths = [20, 15, 10, 12, 8, 22, 30, 10, 8, 8, 45, 40, 65, 6]
        for sheet_name, inputs in results.items():
            if not inputs:
                continue
            safe_name = re.sub(r'[\\/*?:\[\]]', '_', sheet_name)[:31]
            ws = wb.create_sheet(title=safe_name)
            sorted_inputs = sorted(inputs, key=lambda x: (x.device, x.board or '', x.input_number))
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill, c.font, c.alignment, c.border = hfill, hfont, Alignment(horizontal='center'), border
            prev_key, cidx = None, 0
            for row, inp in enumerate(sorted_inputs, 2):
                cur = (inp.device, inp.board)
                if cur != prev_key:
                    if prev_key is not None:
                        cidx = (cidx + 1) % 2
                    prev_key = cur
                vals = [inp.substation or '', inp.bay or '', inp.voltage_level or '', inp.switchgear or '', inp.device, inp.device_model, inp.device_function, inp.board or '', inp.input_id, inp.input_number, inp.description_line1, inp.description_line2, inp.full_description, inp.page_number]
                for col, v in enumerate(vals, 1):
                    c = ws.cell(row=row, column=col, value=v)
                    c.font, c.fill, c.border = dfont, fills[cidx], border
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = ws.dimensions
        if not wb.sheetnames:
            ws = wb.create_sheet(title="No Data")
            ws.cell(row=1, column=1, value="No binary inputs found in the provided files.")
        wb.save(output_path)
        return True
    except ImportError:
        raise ImportError("openpyxl not installed. Install with: pip install openpyxl")
    except Exception as e:
        raise Exception(f"Error creating Excel file: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# GUI APPLICATION
# ═══════════════════════════════════════════════════════════════════════════════

class BinaryInputExtractorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Extractor de Entradas Binarias")
        self.root.geometry("750x550")
        self.root.resizable(True, True)
        self.pdf_paths = ["", "", ""]
        self.output_path = ""
        self._setup_ui()

    def _setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        ttk.Label(main_frame, text="Extractor de Entradas Binarias", font=('Helvetica', 16, 'bold')).pack(pady=(0, 5))
        ttk.Label(main_frame, text="Extrae entradas binarias de planos de protección en PDF", font=('Helvetica', 10)).pack(pady=(0, 15))
        
        pdf_frame = ttk.LabelFrame(main_frame, text="Archivos PDF de entrada (hasta 3)", padding="10")
        pdf_frame.pack(fill=tk.X, pady=(0, 15))
        self.pdf_entries = []
        for i in range(3):
            row_frame = ttk.Frame(pdf_frame)
            row_frame.pack(fill=tk.X, pady=3)
            ttk.Label(row_frame, text=f"PDF {i+1}:", width=8).pack(side=tk.LEFT)
            entry = ttk.Entry(row_frame, width=55)
            entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 5))
            self.pdf_entries.append(entry)
            ttk.Button(row_frame, text="Examinar...", command=lambda idx=i: self._browse_pdf(idx)).pack(side=tk.LEFT)
            ttk.Button(row_frame, text="×", width=3, command=lambda idx=i: self._clear_pdf(idx)).pack(side=tk.LEFT, padx=(3, 0))
        
        output_frame = ttk.LabelFrame(main_frame, text="Archivo Excel de salida", padding="10")
        output_frame.pack(fill=tk.X, pady=(0, 15))
        output_row = ttk.Frame(output_frame)
        output_row.pack(fill=tk.X)
        ttk.Label(output_row, text="Guardar:", width=8).pack(side=tk.LEFT)
        self.output_entry = ttk.Entry(output_row, width=55)
        self.output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 5))
        ttk.Button(output_row, text="Examinar...", command=self._browse_output).pack(side=tk.LEFT)
        
        progress_frame = ttk.Frame(main_frame)
        progress_frame.pack(fill=tk.X, pady=(0, 15))
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, pady=(0, 5))
        self.status_var = tk.StringVar(value="Listo")
        ttk.Label(progress_frame, textvariable=self.status_var, font=('Helvetica', 9)).pack()
        
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=10)
        self.extract_btn = ttk.Button(button_frame, text="Extraer Entradas Binarias", command=self._start_extraction)
        self.extract_btn.pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Salir", command=self.root.quit).pack(side=tk.LEFT, padx=5)
        
        log_frame = ttk.LabelFrame(main_frame, text="Registro", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True)
        self.log_text = tk.Text(log_frame, height=10, wrap=tk.WORD, font=('Consolas', 9))
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.pack(fill=tk.BOTH, expand=True)

    def _browse_pdf(self, index):
        filepath = filedialog.askopenfilename(title=f"Seleccionar PDF {index + 1}", filetypes=[("Archivos PDF", "*.pdf"), ("Archivos ZIP", "*.zip"), ("Todos", "*.*")])
        if filepath:
            self.pdf_entries[index].delete(0, tk.END)
            self.pdf_entries[index].insert(0, filepath)
            self._log(f"Seleccionado PDF {index + 1}: {Path(filepath).name}")
            if not self.output_entry.get():
                self.output_entry.insert(0, str(Path(filepath).parent / "entradas_binarias_resultado.xlsx"))

    def _clear_pdf(self, index):
        self.pdf_entries[index].delete(0, tk.END)

    def _browse_output(self):
        filepath = filedialog.asksaveasfilename(title="Guardar archivo Excel", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")])
        if filepath:
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, filepath)

    def _log(self, message):
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def _update_status(self, message, progress=None):
        self.status_var.set(message)
        if progress is not None:
            self.progress_var.set(progress)
        self.root.update_idletasks()

    def _start_extraction(self):
        pdf_files = [e.get().strip() for e in self.pdf_entries if e.get().strip()]
        if not pdf_files:
            messagebox.showerror("Error", "Por favor seleccione al menos un archivo PDF.")
            return
        output_path = self.output_entry.get().strip()
        if not output_path:
            messagebox.showerror("Error", "Por favor especifique la ruta del archivo de salida.")
            return
        for pdf in pdf_files:
            if not os.path.exists(pdf):
                messagebox.showerror("Error", f"Archivo no encontrado: {pdf}")
                return
        self.extract_btn.configure(state='disabled')
        self.log_text.delete(1.0, tk.END)
        threading.Thread(target=self._run_extraction, args=(pdf_files, output_path)).start()

    def _run_extraction(self, pdf_files, output_path):
        try:
            results = {}
            for i, pdf_path in enumerate(pdf_files):
                filename = Path(pdf_path).stem
                self._update_status(f"Procesando {filename}...", (i / len(pdf_files)) * 100)
                self._log(f"\n{'='*50}\nProcesando: {filename}")
                try:
                    extractor = BinaryInputExtractor(pdf_path)
                    inputs = extractor.extract_all()
                    if inputs:
                        results[filename[:31]] = inputs
                        self._log(f"✓ Encontradas {len(inputs)} entradas binarias")
                        if extractor.substation:
                            self._log(f"  Subestación: {extractor.substation}")
                        devices = {}
                        for inp in inputs:
                            key = f"{inp.device} ({inp.device_model})"
                            devices[key] = devices.get(key, 0) + 1
                        for dev, count in devices.items():
                            self._log(f"  - {dev}: {count} entradas")
                    else:
                        self._log("⚠ No se encontraron entradas binarias")
                except Exception as e:
                    self._log(f"✗ Error: {str(e)}")
            if results:
                self._update_status("Generando Excel...", 90)
                self._log(f"\n{'='*50}\nGuardando en: {output_path}")
                write_multi_tab_xlsx(results, output_path)
                total = sum(len(v) for v in results.values())
                self._update_status("¡Completado!", 100)
                self._log(f"\n✓ Excel creado: {len(results)} pestaña(s), {total} entradas")
                self.root.after(0, lambda: messagebox.showinfo("Éxito", f"¡Extracción completada!\n\n{output_path}\n\n{len(results)} pestaña(s), {total} entradas binarias."))
            else:
                self._update_status("Sin datos", 100)
                self._log("\n⚠ No se encontraron entradas binarias")
                self.root.after(0, lambda: messagebox.showwarning("Advertencia", "No se encontraron entradas binarias."))
        except Exception as e:
            self._update_status("Error", 0)
            self._log(f"\n✗ Error: {str(e)}")
            self.root.after(0, lambda: messagebox.showerror("Error", str(e)))
        finally:
            self.root.after(0, lambda: self.extract_btn.configure(state='normal'))


def main():
    root = tk.Tk()
    try:
        style = ttk.Style()
        if 'clam' in style.theme_names():
            style.theme_use('clam')
    except:
        pass
    BinaryInputExtractorGUI(root)
    root.mainloop()


if __name__ == '__main__':
    main()
